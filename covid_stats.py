"""
Author: Johnny Pham
File: covid_stats.py
Date (Recent Update): 9/9/2020
Description: Web Scraping program to get Corona statistics and outputting results into an .xlsx file
"""

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis
import openpyxl
import os, sys, subprocess
import datetime

FILENAME = 'fsociety.xlsx'
BASE_URL = "https://www.who.int/emergencies/diseases/novel-coronavirus-2019"
wb = openpyxl.load_workbook(FILENAME)
sheet = wb.active

print("Last known row with value:", sheet.max_row)
print("Row that next data will be written in:", sheet.max_row + 1)

previous_max_row = sheet.max_row
current_max_row = sheet.max_row + 1

# Get current date
now = datetime.datetime.now()
today = now.strftime('%m') + '/' + now.strftime('%d') + '/' +  now.strftime('%Y')

# Hard-coded/Default values
sheet['A' + str(current_max_row)] = today

# Web scrape code
options = Options()
options.headless = True
options.add_argument("--window-size=1920,1200")

if sys.platform == "linux":
    driver = webdriver.Chrome(options=options, executable_path= os.path.dirname(os.path.abspath('covid_stats.py')) + "/chromedriver")
else:
    driver = webdriver.Chrome(options=options, executable_path= os.path.dirname(os.path.abspath('covid_stats.py')) + "\\chromedriver.exe")

print('Fetching data...')
driver.get(BASE_URL)
confirmedDeaths = driver.find_element_by_id('confirmedDeaths')

# Number of fatalities
deathsSoFar = int(confirmedDeaths.text.replace(" ",""))
sheet['B' + str(current_max_row)] = deathsSoFar
driver.quit()

# Difference between current day and previous day
if sheet.max_row > 2:
    sheet['C' + str(current_max_row)] = int(sheet['B' + str(current_max_row)].value) - int(sheet['B' + str(previous_max_row)].value)
else:
    sheet['C' + str(current_max_row)] = sheet['B' + str(current_max_row)].value

# Got this from: https://stackoverflow.com/questions/35074473/inserting-a-table-with-openpyxl
# define a table style
mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                    showRowStripes=True)
# create a table
table = openpyxl.worksheet.table.Table(ref='A1:C' + str(current_max_row),
                                    displayName='coronaCasualities',
                                    tableStyleInfo=mediumStyle)
# add the table to the worksheet
# sheet.add_table(table)
# del sheet._tables[0]

if (len(sheet._charts) > 0):
    del sheet._charts[0]

# Make Graph => Referenced: https://openpyxl.readthedocs.io/en/stable/charts/line.html
coronaChart = LineChart()
coronaChart.title = "Corona Death Stats"
coronaChart.style = 13
coronaChart.y_axis.title = sheet["B1"].value

coronaChart.y_axis.crossAx = 500
coronaChart.x_axis = DateAxis(crossAx=100)
coronaChart.x_axis.number_format = 'mm-dd-yy'
coronaChart.x_axis.majorTimeUnit = "days"
coronaChart.x_axis.title = "Date"

data = Reference(sheet, min_col=2, min_row=1, max_col=3, max_row=current_max_row)
coronaChart.add_data(data, titles_from_data=True)
dates = Reference(sheet, min_col=1, min_row=2, max_row=current_max_row)
coronaChart.set_categories(dates)

deathLine = coronaChart.series[0]
deathLine.marker.symbol = "triangle"
deathLine.graphicalProperties.solidFill = "b30505"
deathLine.graphicalProperties.line.solidFill = "b30505" # Marker outline

deathDiffLine = coronaChart.series[1]
deathDiffLine.graphicalProperties.solidFill = "e08802"
deathDiffLine.graphicalProperties.line.solidFill = "e08802"

sheet.add_chart(coronaChart, "F1")

wb.save(filename=FILENAME)

print("Done")

if sys.platform == "linux":
    file = os.path.dirname(os.path.abspath(FILENAME)) + '/' + FILENAME
    subprocess.call(["xdg-open", file])
else:
    file = os.path.dirname(os.path.abspath(FILENAME)) + "\\" + FILENAME
    os.startfile(file)